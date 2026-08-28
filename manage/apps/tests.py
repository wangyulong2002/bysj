"""管理端测试（T2-1/T2-2/T2-3）。

覆盖验收要点（13.6/13.7）：
- 院系/班级/课程/学期 CRUD + 编码唯一；
- 班级指定辅导员/年级/专业/院系；
- 学期 is_current 事务切换（任意时刻仅一个）；
- 教学班唯一约束 (term, class, course) 冲突 → 4091；
- 排课班级冲突/教师冲突拒绝保存（4091）、无冲突通过；
- 非 admin 角色访问 /admin/api/** → 4031。
"""
import threading

from django.contrib.auth import get_user_model
from django.utils import timezone

from rest_framework.test import APIClient, APITestCase
from django.test import TransactionTestCase

from .models import (
    CampusAnnouncement,
    CampusClass,
    CampusCourse,
    CampusCourseOffering,
    CampusCourseSchedule,
    CampusDepartment,
    CampusLeave,
    CampusMessage,
    CampusRagTask,
    CampusScore,
    CampusScoreAudit,
    CampusStudent,
    CampusTeacher,
    CampusTerm,
)

User = get_user_model()


class AdminBaseTestCase(APITestCase):
    """构造管理端测试环境：admin 用户 + 基础数据。"""

    def setUp(self):
        """构造测试环境：创建 admin/teacher/counselor 用户及院系/班级/课程/学期基础数据。"""
        self.admin = User.objects.create_user(
            username="admin_t2", password="pass1234", nick_name="管理员",
            role_code="admin", is_superuser=True, create_time=None,
        )
        self.teacher = User.objects.create_user(
            username="teacher_t2", password="pass1234", nick_name="张老师",
            role_code="teacher",
        )
        self.counselor = User.objects.create_user(
            username="counselor_t2", password="pass1234", nick_name="李导员",
            role_code="counselor",
        )

        self.dept = CampusDepartment.objects.create(
            dept_name="计算机学院", dept_code="CS", del_flag="0"
        )
        self.cls = CampusClass.objects.create(
            class_name="计科2301", class_code="CS2301", grade="2023",
            major="计算机科学与技术", department=self.dept, counselor=self.counselor,
            del_flag="0",
        )
        self.course = CampusCourse.objects.create(
            course_name="高等数学", course_code="MATH101",
            credit="4.0", hours=64, department=self.dept, del_flag="0",
        )
        self.term = CampusTerm.objects.create(
            term_name="2025-2026学年第一学期", start_date="2025-09-01",
            end_date="2026-01-18", total_weeks=20, is_current="1", del_flag="0",
        )
        self.client.force_authenticate(user=self.admin)


class DepartmentTermTestCase(AdminBaseTestCase):
    """T2-1：院系/学期 CRUD + 编码唯一 + is_current 切换。"""

    def test_department_create_and_list(self):
        """院系创建与列表查询成功。"""
        resp = self.client.post("/admin/api/departments", {"dept_name": "外国语学院", "dept_code": "FL"}, format="json")
        self.assertIn(resp.status_code, (200, 201))
        self.assertEqual(resp.json()["code"], 0)
        resp = self.client.get("/admin/api/departments")
        self.assertEqual(resp.json()["code"], 0)
        body = resp.json()["data"]
        results = body.get("results", body)  # DRF 分页时 results 为列表
        names = [d["dept_name"] for d in results]
        self.assertIn("外国语学院", names)

    def test_department_code_unique(self):
        """院系编码重复 → 4091 冲突。"""
        resp = self.client.post("/admin/api/departments", {"dept_name": "重复院系", "dept_code": "CS"}, format="json")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["code"], 4091)

    def test_class_with_counselor_grade_major(self):
        """班级创建携带辅导员/年级/专业/院系，返回关联展示名。"""
        resp = self.client.post(
            "/admin/api/classes",
            {"class_name": "软工2302", "class_code": "SE2302", "grade": "2023",
             "major": "软件工程", "department_id": self.dept.pk, "counselor_id": self.counselor.pk},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["department_id"], self.dept.pk)
        self.assertEqual(data["counselor_id"], self.counselor.pk)
        self.assertEqual(data["department_name"], "计算机学院")
        self.assertEqual(data["counselor_name"], "李导员")

    def test_class_code_unique(self):
        """班级编码重复 → 4091 冲突。"""
        resp = self.client.post(
            "/admin/api/classes",
            {"class_name": "重复班级", "class_code": "CS2301", "grade": "2023"}, format="json",
        )
        self.assertEqual(resp.json()["code"], 4091)

    def test_term_is_current_switch_only_one(self):
        """新建当前学期时旧学期自动置 0，任意时刻仅一个 is_current=1。"""
        resp = self.client.post(
            "/admin/api/terms",
            {"term_name": "2025-2026学年第二学期", "start_date": "2026-02-23",
             "end_date": "2026-07-05", "total_weeks": 19, "is_current": "1"},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        # 任意时刻仅一个 is_current=1
        count = CampusTerm.objects.filter(is_current="1").count()
        self.assertEqual(count, 1)
        self.assertEqual(CampusTerm.objects.filter(is_current="1").first().term_name,
                         "2025-2026学年第二学期")

    def test_term_update_is_current(self):
        """更新学期为当前学期后同样触发 is_current 唯一切换。"""
        term2 = CampusTerm.objects.create(
            term_name="2025-2026学年第二学期", start_date="2026-02-23",
            end_date="2026-07-05", total_weeks=19, is_current="0", del_flag="0",
        )
        resp = self.client.put(
            f"/admin/api/terms/{term2.pk}",
            {"term_name": term2.term_name, "start_date": "2026-02-23",
             "end_date": "2026-07-05", "total_weeks": 19, "is_current": "1"},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        self.assertEqual(CampusTerm.objects.filter(is_current="1").count(), 1)
        self.assertEqual(CampusTerm.objects.get(pk=term2.pk).is_current, "1")


class OfferingTestCase(AdminBaseTestCase):
    """T2-2：教学班 CRUD + 唯一约束 (term, class, course)。"""

    def _create_offering(self):
        """构造教学班（课程+学期+班级+教师）。"""
        return CampusCourseOffering.objects.create(
            course=self.course, term=self.term, class_field=self.cls,
            teacher=self.teacher, del_flag="0",
        )

    def test_offering_create_with_names(self):
        """教学班创建成功并返回关联名称（课程/学期/班级/教师）。"""
        resp = self.client.post(
            "/admin/api/offerings",
            {"course_id": self.course.pk, "term_id": self.term.pk,
             "class_id": self.cls.pk, "teacher_id": self.teacher.pk},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["course_name"], "高等数学")
        self.assertEqual(data["term_name"], "2025-2026学年第一学期")
        self.assertEqual(data["class_name"], "计科2301")
        self.assertEqual(data["teacher_name"], "张老师")

    def test_offering_unique_constraint(self):
        """教学班 (term, class, course) 组合重复 → 4091 冲突。"""
        self._create_offering()
        resp = self.client.post(
            "/admin/api/offerings",
            {"course_id": self.course.pk, "term_id": self.term.pk,
             "class_id": self.cls.pk, "teacher_id": self.teacher.pk},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 4091)

    def test_offering_requires_all_fields(self):
        """教学班缺必填字段 → 4001 参数错误。"""
        resp = self.client.post(
            "/admin/api/offerings", {"course_id": self.course.pk}, format="json",
        )
        self.assertEqual(resp.json()["code"], 4001)


class ScheduleConflictTestCase(AdminBaseTestCase):
    """T2-3：排课冲突校验（班级冲突/教师冲突/无冲突通过）。"""

    def setUp(self):
        """构造排课测试环境：两个同班同教师的教学班。"""
        super().setUp()
        self.offering = CampusCourseOffering.objects.create(
            course=self.course, term=self.term, class_field=self.cls,
            teacher=self.teacher, del_flag="0",
        )
        self.course2 = CampusCourse.objects.create(
            course_name="大学英语", course_code="ENG102",
            credit="2.0", hours=32, department=self.dept, del_flag="0",
        )
        self.offering2 = CampusCourseOffering.objects.create(
            course=self.course2, term=self.term, class_field=self.cls,
            teacher=self.teacher, del_flag="0",
        )

    def _post_schedule(self, offering, **overrides):
        """向 /admin/api/schedules 发起排课创建请求（可覆盖默认字段）。"""
        payload = {
            "offering_id": offering.pk,
            "week_start": 1,
            "week_end": 16,
            "day_of_week": 1,
            "period_start": 1,
            "period_end": 2,
            "location": "A-301",
        }
        payload.update(overrides)
        return self.client.post("/admin/api/schedules", payload, format="json")

    def test_schedule_create_ok(self):
        """无冲突排课创建成功。"""
        resp = self._post_schedule(self.offering)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["code"], 0)
        self.assertEqual(CampusCourseSchedule.objects.count(), 1)

    def test_class_conflict_rejected(self):
        """同班级同一天时间重叠 → 班级冲突 4091。"""
        self._post_schedule(self.offering)
        # 同班级 + 同一天 + 时间重叠 → 班级冲突
        resp = self._post_schedule(self.offering2)
        self.assertEqual(resp.json()["code"], 4091)
        self.assertIn("班级冲突", resp.json()["message"])

    def test_teacher_conflict_rejected(self):
        """另一班级 + 同教师 + 同一天时间重叠 → 教师冲突 4091。"""
        # 另一个班级 + 同教师 + 同一天 + 时间重叠 → 教师冲突
        cls2 = CampusClass.objects.create(
            class_name="计科2302", class_code="CS2302", grade="2023",
            major="计算机科学与技术", department=self.dept, counselor=self.counselor,
            del_flag="0",
        )
        offering3 = CampusCourseOffering.objects.create(
            course=self.course, term=self.term, class_field=cls2,
            teacher=self.teacher, del_flag="0",
        )
        self._post_schedule(self.offering)
        resp = self._post_schedule(offering3)
        self.assertEqual(resp.json()["code"], 4091)
        self.assertIn("教师冲突", resp.json()["message"])

    def test_non_overlap_period_allowed(self):
        """同班同教师但节次不重叠 → 允许创建。"""
        self._post_schedule(self.offering)
        # 同班级同教师，但节次不重叠 → 允许
        resp = self._post_schedule(self.offering2, period_start=3, period_end=4)
        self.assertEqual(resp.json()["code"], 0)
        self.assertEqual(CampusCourseSchedule.objects.count(), 2)

    def test_different_day_allowed(self):
        """同班同教师但星期不同 → 允许创建。"""
        self._post_schedule(self.offering)
        resp = self._post_schedule(self.offering2, day_of_week=2)
        self.assertEqual(resp.json()["code"], 0)

    def test_schedule_update_conflict(self):
        """更新排课为与已有记录重叠的节次 → 4091 冲突。"""
        CampusCourseSchedule.objects.create(
            offering=self.offering, week_start=1, week_end=16, day_of_week=1,
            period_start=1, period_end=2, location="A-301", del_flag="0",
        )
        other = CampusCourseSchedule.objects.create(
            offering=self.offering2, week_start=1, week_end=16, day_of_week=1,
            period_start=3, period_end=4, location="A-302", del_flag="0",
        )
        # 把 other 改成与同班级同教师 A 重叠的节次 → 拒绝
        resp = self.client.put(
            f"/admin/api/schedules/{other.pk}",
            {"offering_id": self.offering2.pk, "week_start": 1, "week_end": 16,
             "day_of_week": 1, "period_start": 1, "period_end": 2, "location": "B-201"},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 4091)


class AutoCodeGenerateTestCase(AdminBaseTestCase):
    """编码自动生成：院系/班级/课程/学生/教师 编码留空时自动生成唯一值。"""

    def test_auto_generated_codes(self):
        """各业务编码留空提交，自动生成对应前缀唯一编码；学生/教师账号联动。"""
        # 院系
        resp = self.client.post("/admin/api/departments", {"dept_name": "自动院系"}, format="json")
        self.assertEqual(resp.json()["code"], 0)
        dept_code = resp.json()["data"]["dept_code"]
        self.assertTrue(dept_code.startswith("DEPT"))
        # 班级
        resp = self.client.post(
            "/admin/api/classes",
            {"class_name": "自动班", "grade": "2026", "department_id": self.dept.pk},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        self.assertTrue(resp.json()["data"]["class_code"].startswith("CLS"))
        # 课程
        resp = self.client.post(
            "/admin/api/courses",
            {"course_name": "自动课", "credit": 2.0, "hours": 32, "department_id": self.dept.pk},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        self.assertTrue(resp.json()["data"]["course_code"].startswith("CRS"))
        # 学生（自动学号 + 联动账号 username=学号）
        resp = self.client.post(
            "/admin/api/students",
            {"nick_name": "自动生", "class_id": self.cls.pk, "enroll_year": "2026", "password": "123456"},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        student_no = resp.json()["data"]["student_no"]
        self.assertTrue(student_no.startswith("S"))
        self.assertTrue(User.objects.filter(username=student_no, role_code="student").exists())
        # 教师（自动工号 + 联动账号 username=工号）
        resp = self.client.post(
            "/admin/api/teachers",
            {"nick_name": "自动师", "department_id": self.dept.pk, "title": "讲师", "password": "123456"},
            format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        teacher_no = resp.json()["data"]["teacher_no"]
        self.assertTrue(teacher_no.startswith("T"))
        self.assertTrue(User.objects.filter(username=teacher_no, role_code="teacher").exists())

    def test_auto_code_unique_and_update_keep(self):
        """两次创建编码不重复；更新时留空保留原编码。"""
        resp1 = self.client.post("/admin/api/departments", {"dept_name": "自动院系A"}, format="json")
        resp2 = self.client.post("/admin/api/departments", {"dept_name": "自动院系B"}, format="json")
        code1 = resp1.json()["data"]["dept_code"]
        code2 = resp2.json()["data"]["dept_code"]
        self.assertNotEqual(code1, code2)
        # 更新留空 → 保留原编码
        pk = resp1.json()["data"]["id"]
        resp3 = self.client.put(f"/admin/api/departments/{pk}", {"dept_name": "改名院系"}, format="json")
        self.assertEqual(resp3.json()["code"], 0)
        self.assertEqual(resp3.json()["data"]["dept_code"], code1)


class AdminPermissionTestCase(AdminBaseTestCase):
    """T2-1 配置：非 admin 角色访问 /admin/api/** → 4031。"""

    def test_non_admin_denied(self):
        """非 admin 角色访问 /admin/api/** → 4031。"""
        self.client.force_authenticate(user=self.counselor)
        resp = self.client.get("/admin/api/departments")
        self.assertEqual(resp.json()["code"], 4031)
        resp = self.client.get("/admin/api/schedules")
        self.assertEqual(resp.json()["code"], 4031)


class ScheduleConcurrentTestCase(TransactionTestCase):
    """T2-3 并发（B-04/P1-13，验收标准 17）：并发保存重叠时段仅一方成功。

    使用 TransactionTestCase（真实提交事务），两线程同时 POST 同班级/同教师
    同一天重叠节次的排课，验证 FOR UPDATE 锁 + 冲突校验在并发下仍生效。
    """

    def setUp(self):
        """构造并发测试环境：两个同班同教师的教学班。"""
        self.admin = User.objects.create_user(
            username="admin_t2c", password="pass1234", nick_name="管理员",
            role_code="admin", is_superuser=True,
        )
        self.teacher = User.objects.create_user(
            username="teacher_t2c", password="pass1234", nick_name="张老师",
            role_code="teacher",
        )
        self.dept = CampusDepartment.objects.create(dept_name="计算机学院", dept_code="CSC", del_flag="0")
        self.cls = CampusClass.objects.create(
            class_name="计科2301", class_code="CS2301C", grade="2023",
            major="计算机科学与技术", department=self.dept, del_flag="0",
        )
        self.term = CampusTerm.objects.create(
            term_name="2025-2026学年第一学期", start_date="2025-09-01",
            end_date="2026-01-18", total_weeks=20, is_current="1", del_flag="0",
        )
        self.course1 = CampusCourse.objects.create(
            course_name="高等数学", course_code="MATH101C",
            credit="4.0", hours=64, department=self.dept, del_flag="0",
        )
        self.course2 = CampusCourse.objects.create(
            course_name="大学英语", course_code="ENG102C",
            credit="2.0", hours=32, department=self.dept, del_flag="0",
        )
        self.offering1 = CampusCourseOffering.objects.create(
            course=self.course1, term=self.term, class_field=self.cls,
            teacher=self.teacher, del_flag="0",
        )
        self.offering2 = CampusCourseOffering.objects.create(
            course=self.course2, term=self.term, class_field=self.cls,
            teacher=self.teacher, del_flag="0",
        )

    def test_concurrent_overlap_only_one_succeeds(self):
        """两线程并发提交重叠排课：FOR UPDATE 锁下仅一方成功、另一方 4091。"""
        results = {}
        barrier = threading.Barrier(2)

        def _post(offering_id, key):
            """线程内发起排课请求，记录响应 code。"""
            client = APIClient()
            client.force_authenticate(user=self.admin)
            barrier.wait()  # 尽量同时发起
            resp = client.post(
                "/admin/api/schedules",
                {"offering_id": offering_id, "week_start": 1, "week_end": 16,
                 "day_of_week": 1, "period_start": 1, "period_end": 2, "location": "A-301"},
                format="json",
            )
            results[key] = resp.json()["code"]

        threads = [
            threading.Thread(target=_post, args=(self.offering1.pk, "a")),
            threading.Thread(target=_post, args=(self.offering2.pk, "b")),
        ]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        self.assertEqual(sorted(results.values()), [0, 4091], f"并发结果应为一方成功一方冲突，实际 {results}")
        self.assertEqual(CampusCourseSchedule.objects.count(), 1)


class AnnouncementTestCase(AdminBaseTestCase):
    """T3-1：公告 CRUD + 状态流转（草稿→发布→下架）+ RAG 任务联动 + 权限。

    覆盖验收要点（验收标准 3）：
    - 类型（校园/院系/班级）与单目标校验（4.2 P1-07）；
    - 置顶、状态流转闭环；
    - 发布记录 publish_time/publisher_id（T3-1 完成事项）；
    - 发布/下架写 campus_rag_task（8.3 触发点）；
    - 非 admin 角色访问 → 4031。
    """

    def setUp(self):
        """构造公告测试环境：admin 用户 + 院系/班级基础数据。"""
        super().setUp()

    def _create_draft(self, **overrides):
        """创建草稿公告（默认校园公告）。"""
        payload = {
            "title": "测试公告", "content": "公告内容", "ann_type": "1",
            "is_top": "0", "status": "0",
        }
        payload.update(overrides)
        return self.client.post("/admin/api/announcements", payload, format="json")

    def _rag_tasks(self, ann_pk, operation):
        """查询该公告的 RAG 任务（source_type=1 公告）。"""
        return CampusRagTask.objects.filter(
            source_type="1", source_id=ann_pk, operation=operation, del_flag="0"
        )

    def test_announcement_create_draft(self):
        """创建草稿公告：status=0、无 publish_time。"""
        resp = self._create_draft()
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["status"], "0")
        self.assertEqual(data["ann_type"], "1")
        self.assertIsNone(data["publish_time"])

    def test_announcement_create_class_requires_target(self):
        """班级公告（ann_type=3）未选目标班级 → 4001。"""
        resp = self._create_draft(ann_type="3")
        self.assertEqual(resp.json()["code"], 4001)
        self.assertIn("班级", resp.json()["message"])

    def test_announcement_create_department_requires_target(self):
        """院系公告（ann_type=2）未选目标院系 → 4001。"""
        resp = self._create_draft(ann_type="2")
        self.assertEqual(resp.json()["code"], 4001)
        self.assertIn("院系", resp.json()["message"])

    def test_announcement_create_school_ignores_target(self):
        """校园公告指定目标班级/院系 → 4001。"""
        resp = self._create_draft(ann_type="1", target_class_id=self.cls.pk)
        self.assertEqual(resp.json()["code"], 4001)

    def test_publish_sets_time_and_writes_rag_task(self):
        """发布：status=1、记录 publish_time/publisher、写 RAG upsert 任务（T3-1）。"""
        resp = self._create_draft()
        ann_pk = resp.json()["data"]["id"]
        resp = self.client.post(f"/admin/api/announcements/{ann_pk}/publish")
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["status"], "1")
        self.assertIsNotNone(data["publish_time"])
        self.assertEqual(data["publisher_id"], self.admin.pk)
        # RAG upsert 任务（operation=1, source_type=1 公告）
        self.assertTrue(self._rag_tasks(ann_pk, "1").exists())
        self.assertEqual(self._rag_tasks(ann_pk, "1").count(), 1)

    def test_take_down_writes_delete_rag_task(self):
        """下架：status=2、写 RAG delete 任务。"""
        resp = self._create_draft()
        ann_pk = resp.json()["data"]["id"]
        self.client.post(f"/admin/api/announcements/{ann_pk}/publish")
        resp = self.client.post(f"/admin/api/announcements/{ann_pk}/take-down")
        self.assertEqual(resp.json()["code"], 0)
        self.assertEqual(resp.json()["data"]["status"], "2")
        self.assertTrue(self._rag_tasks(ann_pk, "2").exists())

    def test_status_flow_validation(self):
        """状态流转校验：未发布不可下架、已发布不可重复发布。"""
        resp = self._create_draft()
        ann_pk = resp.json()["data"]["id"]
        resp = self.client.post(f"/admin/api/announcements/{ann_pk}/take-down")
        self.assertEqual(resp.json()["code"], 4001)
        self.client.post(f"/admin/api/announcements/{ann_pk}/publish")
        resp = self.client.post(f"/admin/api/announcements/{ann_pk}/publish")
        self.assertEqual(resp.json()["code"], 4001)

    def test_republish_after_take_down(self):
        """下架后重新发布：再次写 upsert 任务，publish_time 保持首次发布值。"""
        resp = self._create_draft()
        ann_pk = resp.json()["data"]["id"]
        self.client.post(f"/admin/api/announcements/{ann_pk}/publish")
        first_time = CampusAnnouncement.objects.get(pk=ann_pk).publish_time
        self.client.post(f"/admin/api/announcements/{ann_pk}/take-down")
        resp = self.client.post(f"/admin/api/announcements/{ann_pk}/publish")
        self.assertEqual(resp.json()["code"], 0)
        ann = CampusAnnouncement.objects.get(pk=ann_pk)
        self.assertEqual(ann.publish_time, first_time)  # 首次发布时间不回退
        self.assertEqual(self._rag_tasks(ann_pk, "1").count(), 2)  # 两次发布两次 upsert

    def test_delete_published_writes_rag_task(self):
        """删除已发布公告：写 delete RAG 任务 + 逻辑删除。"""
        resp = self._create_draft()
        ann_pk = resp.json()["data"]["id"]
        self.client.post(f"/admin/api/announcements/{ann_pk}/publish")
        resp = self.client.delete(f"/admin/api/announcements/{ann_pk}")
        self.assertEqual(resp.json()["code"], 0)
        ann = CampusAnnouncement.objects.get(pk=ann_pk)
        self.assertEqual(ann.del_flag, "2")  # 逻辑删除
        self.assertTrue(self._rag_tasks(ann_pk, "2").exists())

    def test_announcement_non_admin_denied(self):
        """非 admin 角色访问公告管理接口 → 4031。"""
        self.client.force_authenticate(user=self.counselor)
        resp = self.client.get("/admin/api/announcements")
        self.assertEqual(resp.json()["code"], 4031)


class StudentTeacherTestCase(AdminBaseTestCase):
    """T2-8/T2-9：学生/教师档案管理（方案 B：档案 ↔ sys_user 账号联动）。

    覆盖：
    - 创建档案自动创建 sys_user（角色/登录名/姓名/初始密码）；
    - 学号/工号唯一冲突 → 4091；账号已存在 → 4001；
    - 更新同步姓名/登录名；删除（逻辑）联动停用账号；
    - 非 admin 4031。
    """

    def _create_student(self, **overrides):
        """创建学生档案（默认：学号 S20260001/姓名 测试学生/班级 self.cls）。"""
        payload = {
            "student_no": "S20260001", "nick_name": "测试学生",
            "class_id": self.cls.pk, "enroll_year": "2026", "password": "",
        }
        payload.update(overrides)
        return self.client.post("/admin/api/students", payload, format="json")

    def test_student_create_creates_account(self):
        """创建学生档案 → 自动创建 sys_user（role=student，username=学号，姓名同步）。"""
        resp = self._create_student()
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["student_no"], "S20260001")
        self.assertEqual(data["username"], "S20260001")
        self.assertEqual(data["class_name"], self.cls.class_name)
        user = User.objects.get(username="S20260001")
        self.assertEqual(user.role_code, "student")
        self.assertEqual(user.nick_name, "测试学生")
        self.assertEqual(user.status, "0")
        self.assertEqual(user.student_no, "S20260001")
        # 档案 user_id 指向该账号
        self.assertEqual(CampusStudent.objects.get(pk=data["id"]).user_id, user.pk)

    def test_student_duplicate_no_conflict(self):
        """学号重复 → 4091（数据库唯一兜底）。"""
        self._create_student()
        resp = self._create_student(student_no="S20260001")
        self.assertEqual(resp.json()["code"], 4091)

    def test_student_account_exists_rejected(self):
        """登录账号已存在（学号被占用）→ 4001 明确提示。"""
        # 预置同名登录账号（不经过接口），再通过接口创建 → 校验拒绝
        User.objects.create_user(username="S20260001", password="x",
                                 role_code="student", del_flag="0",
                                 nick_name="已存在")
        resp = self._create_student(student_no="S20260001")
        self.assertEqual(resp.json()["code"], 4001)
        self.assertIn("已存在", resp.json()["message"])

    def test_student_update_syncs_account(self):
        """更新档案：姓名/学号变更同步 sys_user。"""
        resp = self._create_student()
        sid = resp.json()["data"]["id"]
        resp = self.client.put(f"/admin/api/students/{sid}",
                               {"student_no": "S20260001", "nick_name": "新名字",
                                "class_id": self.cls.pk, "enroll_year": "2026"},
                               format="json")
        self.assertEqual(resp.json()["code"], 0)
        user = User.objects.get(username="S20260001")
        self.assertEqual(user.nick_name, "新名字")
        # 学号变更 → 登录名同步
        resp = self.client.put(f"/admin/api/students/{sid}",
                               {"student_no": "S20260002", "nick_name": "新名字",
                                "class_id": self.cls.pk, "enroll_year": "2026"},
                               format="json")
        self.assertEqual(resp.json()["code"], 0)
        self.assertTrue(User.objects.filter(username="S20260002").exists())
        self.assertFalse(User.objects.filter(username="S20260001").exists())

    def test_student_delete_disables_account(self):
        """删除学生档案（逻辑）→ 关联账号停用（del_flag=2 + status=1）。"""
        resp = self._create_student()
        sid = resp.json()["data"]["id"]
        resp = self.client.delete(f"/admin/api/students/{sid}")
        self.assertEqual(resp.json()["code"], 0)
        user = User.objects.get(username="S20260001")
        self.assertEqual(user.del_flag, "2")
        self.assertEqual(user.status, "1")

    def test_teacher_create_creates_account(self):
        """创建教师档案 → 自动创建 sys_user（role=teacher，username=工号）。"""
        payload = {"teacher_no": "T20260001", "nick_name": "测试教师",
                   "title": "副教授", "department_id": self.dept.pk, "password": ""}
        resp = self.client.post("/admin/api/teachers", payload, format="json")
        self.assertEqual(resp.json()["code"], 0)
        user = User.objects.get(username="T20260001")
        self.assertEqual(user.role_code, "teacher")
        self.assertEqual(user.nick_name, "测试教师")
        self.assertEqual(user.teacher_no, "T20260001")

    def test_student_teacher_non_admin_denied(self):
        """非 admin 访问学生/教师管理 → 4031。"""
        self.client.force_authenticate(user=self.counselor)
        self.assertEqual(self.client.get("/admin/api/students").json()["code"], 4031)
        self.assertEqual(self.client.get("/admin/api/teachers").json()["code"], 4031)


class TeacherCounselorTestCase(AdminBaseTestCase):
    """T2-9 兼任辅导员（ADR-010）：教师可被指定为班级辅导员（≤2 班，前后端校验）。

    覆盖：
    - 创建教师兼任 1~2 班 → 班级 counselor_id 写入；
    - 兼任 >2 班 → 4001；兼任已有辅导员的班级 → 4001；
    - 更新：改兼任班级（旧清空新写入）/ 取消兼任（全部清空）；
    - 删除教师 → 兼任班级 counselor_id 清空；
    - 无辅导员班级选项接口（without_counselor=1）。
    """

    def setUp(self):
        """复用基类数据 + 新建一个无辅导员班级（self.cls 已指定辅导员李导员）。"""
        super().setUp()
        self.cls_free = CampusClass.objects.create(
            class_name="软工2301", class_code="CS2301B", grade="2023",
            major="软件工程", department=self.dept, del_flag="0",
        )

    def _create_teacher(self, teacher_no="T9-001", is_counselor=False, class_ids=None):
        """创建教师（可带兼任参数）。"""
        payload = {"teacher_no": teacher_no, "nick_name": "兼任测试",
                   "title": "讲师", "department_id": self.dept.pk, "password": ""}
        if is_counselor is not False:
            payload["is_counselor"] = is_counselor
        if class_ids is not None:
            payload["counselor_class_ids"] = class_ids
        return self.client.post("/admin/api/teachers", payload, format="json")

    def test_teacher_create_with_counselor(self):
        """创建教师兼任 1 个班级 → campus_class.counselor_id 写入该教师。"""
        resp = self._create_teacher(is_counselor=True, class_ids=[self.cls_free.pk])
        self.assertEqual(resp.json()["code"], 0)
        cls = CampusClass.objects.get(pk=self.cls_free.pk)
        user = User.objects.get(username="T9-001")
        self.assertEqual(cls.counselor_id, user.pk)
        data = resp.json()["data"]
        self.assertEqual(data["counselor_class_id_list"], [self.cls_free.pk])
        self.assertEqual(data["counselor_class_names"], [self.cls_free.class_name])

    def test_teacher_counselor_max_two(self):
        """兼任超过 2 个班级 → 4001（前后端限制，ADR-010）。"""
        cls2 = CampusClass.objects.create(class_name="测试班2", class_code="T92",
                                          grade="2024", major="测试", del_flag="0")
        cls3 = CampusClass.objects.create(class_name="测试班3", class_code="T93",
                                          grade="2024", major="测试", del_flag="0")
        resp = self._create_teacher(is_counselor=True,
                                    class_ids=[self.cls_free.pk, cls2.pk, cls3.pk])
        self.assertEqual(resp.json()["code"], 4001)
        self.assertIn("最多兼任 2 个", resp.json()["message"])

    def test_teacher_counselor_occupied_class(self):
        """兼任已有辅导员的班级 → 4001（占用校验）。"""
        # 先建一个辅导员并指定给 self.cls
        counselor = User.objects.create_user(username="T9-own", password="x",
                                             role_code="counselor", del_flag="0")
        CampusClass.objects.filter(pk=self.cls.pk).update(counselor_id=counselor.pk)
        resp = self._create_teacher(is_counselor=True, class_ids=[self.cls.pk])
        self.assertEqual(resp.json()["code"], 4001)
        self.assertIn("已有辅导员", resp.json()["message"])

    def test_teacher_update_switch_counselor_classes(self):
        """更新：改兼任班级 → 旧班级清空、新班级写入。"""
        cls2 = CampusClass.objects.create(class_name="测试班B", class_code="T9B",
                                          grade="2024", major="测试", del_flag="0")
        resp = self._create_teacher(is_counselor=True, class_ids=[self.cls_free.pk])
        tid = resp.json()["data"]["id"]
        user = User.objects.get(username="T9-001")
        resp = self.client.put(f"/admin/api/teachers/{tid}",
                               {"teacher_no": "T9-001", "nick_name": "兼任测试",
                                "title": "讲师", "department_id": self.dept.pk,
                                "is_counselor": True, "counselor_class_ids": [cls2.pk]},
                               format="json")
        self.assertEqual(resp.json()["code"], 0)
        self.assertEqual(CampusClass.objects.get(pk=self.cls_free.pk).counselor_id, None)
        self.assertEqual(CampusClass.objects.get(pk=cls2.pk).counselor_id, user.pk)

    def test_teacher_update_cancel_counselor(self):
        """更新：取消兼任（is_counselor=False）→ 兼任班级全部清空。"""
        resp = self._create_teacher(is_counselor=True, class_ids=[self.cls_free.pk])
        tid = resp.json()["data"]["id"]
        resp = self.client.put(f"/admin/api/teachers/{tid}",
                               {"teacher_no": "T9-001", "nick_name": "兼任测试",
                                "title": "讲师", "department_id": self.dept.pk,
                                "is_counselor": False, "counselor_class_ids": []},
                               format="json")
        self.assertEqual(resp.json()["code"], 0)
        self.assertEqual(CampusClass.objects.get(pk=self.cls_free.pk).counselor_id, None)

    def test_teacher_delete_clears_counselor(self):
        """删除教师 → 兼任班级 counselor_id 清空 + 账号停用。"""
        self._create_teacher(is_counselor=True, class_ids=[self.cls_free.pk])
        tid = CampusTeacher.objects.get(user__username="T9-001").pk
        resp = self.client.delete(f"/admin/api/teachers/{tid}")
        self.assertEqual(resp.json()["code"], 0)
        self.assertEqual(CampusClass.objects.get(pk=self.cls_free.pk).counselor_id, None)
        self.assertEqual(User.objects.get(username="T9-001").del_flag, "2")

    def test_class_options_without_counselor(self):
        """无辅导员班级选项接口：仅返回 counselor_id IS NULL 的班级。"""
        # self.cls 已指定辅导员（李导员）→ 不在列表；cls_free 无辅导员 → 在列表
        counselor = User.objects.create_user(username="T9-own2", password="x",
                                             role_code="counselor", del_flag="0")
        cls_occ = CampusClass.objects.create(class_name="已占用班", class_code="T9O",
                                             grade="2024", major="测试",
                                             counselor_id=counselor.pk, del_flag="0")
        resp = self.client.get("/admin/api/classes/options?without_counselor=1")
        self.assertEqual(resp.json()["code"], 0)
        ids = [c["id"] for c in resp.json()["data"]]
        self.assertIn(self.cls_free.pk, ids)
        self.assertNotIn(self.cls.pk, ids)
        self.assertNotIn(cls_occ.pk, ids)


class ScoreAdminTestCase(AdminBaseTestCase):
    """T4-1：成绩管理（发布/撤销发布/审计 + Excel 导入）。"""

    def setUp(self):
        """构造：学生（self.cls 班）+ 教学班 + 成绩记录。"""
        super().setUp()
        self.course = CampusCourse.objects.create(
            course_name="测试课程M4", course_code="M4K", credit=3.0,
            hours=48, department=self.dept, del_flag="0",
        )
        self.t_user = User.objects.create_user(username="M4T001", password="x",
                                               role_code="teacher", del_flag="0")
        self.teacher = CampusTeacher.objects.create(
            user=self.t_user, teacher_no="M4T001", title="讲师",
            department=self.dept, del_flag="0",
        )
        s_user = User.objects.create_user(username="M4S001", password="x",
                                          role_code="student", del_flag="0")
        self.student = CampusStudent.objects.create(
            user=s_user, student_no="M4S001", class_field=self.cls,
            enroll_year="2023", del_flag="0",
        )
        self.term2 = CampusTerm.objects.create(
            term_name="2025-2026学年第二学期", start_date="2026-02-01",
            end_date="2026-07-01", total_weeks=20, is_current="0", del_flag="0",
        )
        self.offering2 = CampusCourseOffering.objects.create(
            course=self.course, term=self.term2, class_field=self.cls,
            teacher=self.t_user, del_flag="0",
        )
        self.score = CampusScore.objects.create(
            student=self.student, offering=self.offering2,
            usual_score="80.00", exam_score="70.00", total_score="74.00",
            usual_ratio=40, exam_ratio=60, is_published="0", version=0,
            del_flag="0",
        )

    def test_score_list_and_publish(self):
        """成绩列表可见 + 发布 → is_published=1 + 审计(operation=3)。"""
        resp = self.client.get(f"/admin/api/scores?offering_id={self.offering2.pk}")
        self.assertEqual(resp.json()["code"], 0)
        resp = self.client.post(f"/admin/api/scores/{self.score.pk}/publish")
        self.assertEqual(resp.json()["code"], 0)
        self.score.refresh_from_db()
        self.assertEqual(self.score.is_published, "1")
        self.assertIsNotNone(self.score.publish_by)
        self.assertTrue(CampusScoreAudit.objects.filter(
            student=self.student, offering=self.offering2, operation="3"
        ).exists())

    def test_score_unpublish_writes_audit(self):
        """撤销发布 → is_published=0 + 审计(operation=4)。"""
        self.client.post(f"/admin/api/scores/{self.score.pk}/publish")
        resp = self.client.post(f"/admin/api/scores/{self.score.pk}/unpublish")
        self.assertEqual(resp.json()["code"], 0)
        self.score.refresh_from_db()
        self.assertEqual(self.score.is_published, "0")
        self.assertTrue(CampusScoreAudit.objects.filter(
            student=self.student, offering=self.offering2, operation="4"
        ).exists())

    def test_score_audit_list(self):
        """审计查询接口可查。"""
        self.client.post(f"/admin/api/scores/{self.score.pk}/publish")
        resp = self.client.get("/admin/api/score-audits",
                               {"student_id": self.student.pk})
        self.assertEqual(resp.json()["code"], 0)
        self.assertGreaterEqual(len(resp.json()["data"]["results"]), 1)

    def test_score_import_excel(self):
        """Excel 导入：正确学号导入成功，错误学号给出错误行提示。"""
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["学号", "平时成绩", "考试成绩"])
        ws.append([self.student.student_no, 88, 92])
        ws.append(["NOT_EXIST", 80, 70])  # 错误学号
        ws.append([self.student.student_no, 150, 70])  # 超范围
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = self.client.post(
            f"/admin/api/scores/import?offering_id={self.offering2.pk}",
            {"file": buf}, format="multipart",
        )
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["imported"], 1)
        self.assertEqual(len(data["errors"]), 2)

    def test_score_import_excel_new_template(self):
        """新模板（班级|学科|学号|平时成绩|考试成绩）：按 班级+学科 定位当前学期教学班。"""
        from io import BytesIO

        from openpyxl import Workbook

        # 当前学期（基类 self.term is_current=1）下的教学班
        offering_cur = CampusCourseOffering.objects.create(
            course=self.course, term=self.term, class_field=self.cls,
            teacher=self.t_user, del_flag="0",
        )
        wb = Workbook()
        ws = wb.active
        ws.append(["班级", "学科", "学号", "平时成绩", "考试成绩"])
        ws.append([self.cls.class_name, self.course.course_name, self.student.student_no, 88, 92])  # 成功
        ws.append([self.cls.class_name, self.course.course_name, self.student.student_no, 150, 70])  # 超范围
        ws.append(["不存在的班级", self.course.course_name, self.student.student_no, 80, 70])  # 班级不存在
        ws.append([self.cls.class_name, "不存在的课程", self.student.student_no, 80, 70])  # 学科不存在
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = self.client.post("/admin/api/scores/import", {"file": buf}, format="multipart")
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["imported"], 1)
        self.assertEqual(len(data["errors"]), 3)
        self.assertTrue(CampusScore.objects.filter(
            student=self.student, offering=offering_cur, total_score="90.40"
        ).exists())

    def test_score_import_template_download(self):
        """模板下载：GET /admin/api/scores/import-template 返回 xlsx，表头六列，自动填充班级学生学号。"""
        resp = self.client.get(
            f"/admin/api/scores/import-template?class_id={self.cls.pk}&course_id={self.course.pk}"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        header = [c.value for c in ws[1]][:5]
        self.assertEqual(header, ["班级", "学科", "学号", "平时成绩", "考试成绩"])
        # 数据行预填 班级/学科/学号（自动拉取该班级学生）
        self.assertEqual(ws.cell(row=2, column=1).value, self.cls.class_name)
        self.assertEqual(ws.cell(row=2, column=2).value, self.course.course_name)
        self.assertEqual(ws.cell(row=2, column=3).value, self.student.student_no)

    def test_score_batch_publish(self):
        """批量发布：未发布→已发布+审计(3)，已发布跳过。"""
        s2_user = User.objects.create_user(username="M4S002", password="x",
                                           role_code="student", del_flag="0")
        student2 = CampusStudent.objects.create(
            user=s2_user, student_no="M4S002", class_field=self.cls,
            enroll_year="2023", del_flag="0",
        )
        s2 = CampusScore.objects.create(
            student=student2, offering=self.offering2,
            usual_score="90.00", exam_score="80.00", total_score="84.00",
            usual_ratio=40, exam_ratio=60, is_published="1", version=0, del_flag="0",
        )
        resp = self.client.post(
            "/admin/api/scores/batch-publish",
            {"ids": [self.score.pk, s2.pk]}, format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["published"], 1)
        self.assertEqual(data["skipped"], 1)
        self.score.refresh_from_db()
        self.assertEqual(self.score.is_published, "1")
        self.assertTrue(CampusScoreAudit.objects.filter(
            student=self.student, offering=self.offering2, operation="3"
        ).exists())

    def test_score_batch_unpublish(self):
        """批量撤销：已发布→未发布+审计(4)，未发布跳过。"""
        self.client.post(f"/admin/api/scores/{self.score.pk}/publish")
        s2_user = User.objects.create_user(username="M4S003", password="x",
                                           role_code="student", del_flag="0")
        student2 = CampusStudent.objects.create(
            user=s2_user, student_no="M4S003", class_field=self.cls,
            enroll_year="2023", del_flag="0",
        )
        s2 = CampusScore.objects.create(
            student=student2, offering=self.offering2,
            usual_score="90.00", exam_score="80.00", total_score="84.00",
            usual_ratio=40, exam_ratio=60, is_published="0", version=0, del_flag="0",
        )
        resp = self.client.post(
            "/admin/api/scores/batch-unpublish",
            {"ids": [self.score.pk, s2.pk]}, format="json",
        )
        self.assertEqual(resp.json()["code"], 0)
        data = resp.json()["data"]
        self.assertEqual(data["unpublished"], 1)
        self.assertEqual(data["skipped"], 1)
        self.score.refresh_from_db()
        self.assertEqual(self.score.is_published, "0")
        self.assertTrue(CampusScoreAudit.objects.filter(
            student=self.student, offering=self.offering2, operation="4"
        ).exists())


class LeaveAdminTestCase(AdminBaseTestCase):
    """T5-6：请假管理（列表 + 管理员干预 P1-15）。"""

    def setUp(self):
        """构造：学生 + 请假记录。"""
        super().setUp()
        self.stu_user = User.objects.create_user(username="M5S001", password="x",
                                                 role_code="student", del_flag="0")
        self.student = CampusStudent.objects.create(
            user=self.stu_user, student_no="M5S001", class_field=self.cls,
            enroll_year="2023", del_flag="0",
        )
        self.leave = CampusLeave.objects.create(
            student=self.student, leave_type="1", reason="测试请假",
            start_time=timezone.now(), end_time=timezone.now() + timezone.timedelta(hours=4),
            leave_duration_minutes=240, total_days="0.2", status="0", version=0,
            create_time=timezone.now(), del_flag="0",
        )

    def test_leave_list_and_intervene(self):
        """请假列表 + 干预（0→1，reason 必填，消息生成）。"""
        resp = self.client.get("/admin/api/leaves")
        self.assertEqual(resp.json()["code"], 0)
        # 缺 reason → 4001
        resp = self.client.post(f"/admin/api/leaves/{self.leave.pk}/intervene",
                                {"status": "1"}, format="json")
        self.assertEqual(resp.json()["code"], 4001)
        # 干预 0→1
        resp = self.client.post(f"/admin/api/leaves/{self.leave.pk}/intervene",
                                {"status": "1", "reason": "核实情况属实"}, format="json")
        self.assertEqual(resp.json()["code"], 0)
        self.leave.refresh_from_db()
        self.assertEqual(self.leave.status, "1")
        self.assertTrue(CampusMessage.objects.filter(
            user_id=self.stu_user.pk, business_type="leave",
            business_id=self.leave.pk, msg_type="2",
        ).exists())
        # 终态回退（1→0）也需 reason
        resp = self.client.post(f"/admin/api/leaves/{self.leave.pk}/intervene",
                                {"status": "0"}, format="json")
        self.assertEqual(resp.json()["code"], 4001)
        resp = self.client.post(f"/admin/api/leaves/{self.leave.pk}/intervene",
                                {"status": "0", "reason": "系统误操作回退"}, format="json")
        self.assertEqual(resp.json()["code"], 0)

    def test_leave_non_admin_denied(self):
        """非 admin 访问请假管理 → 4031。"""
        self.client.force_authenticate(user=self.counselor)
        self.assertEqual(self.client.get("/admin/api/leaves").json()["code"], 4031)
