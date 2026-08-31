"""管理端 DRF 视图集（T2-1/T2-2/T2-3）。

设计基线：v2.2 4.1（排课冲突）、5.1（is_current 唯一/逻辑删除/事务）、6.4（/admin/api/**）。
- 统一权限：仅 admin 角色可访问（DRF 权限类，6.4/P1-10）。
- 删除一律逻辑删除（del_flag='2'，5.1，不物理删除被引用记录）。
- 学期 is_current 切换：事务内先将旧学期置 0，再置新学期为 1（5.1/P1-04）。
- 教学班唯一约束：(term, class, course) 数据库兜底 + 友好提示（T2-2）。
- 排课冲突（4.1 P0-06/B-04/P1-13）：班级冲突 + 教师冲突；
  冲突校验与写入同事务，固定 MySQL `SELECT ... FOR UPDATE`，
  按 class_id/teacher_id 升序锁行，补死锁重试（OperationalError 1213）。
"""
from django.contrib.auth import get_user_model
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.db.utils import OperationalError
from django.utils import timezone

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.filters import SearchFilter
from rest_framework.exceptions import (
    ErrorDetail,
    NotAuthenticated,
    NotFound,
    PermissionDenied,
    ValidationError,
)
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.views import exception_handler as drf_exception_handler

from .announcement_cache import get_redis_client
from .announcement_flow import on_announcement_deleted, on_announcement_saved
from .knowledge_flow import (
    compute_content_hash,
    on_knowledge_deleted,
    on_knowledge_saved,
)
from .models import (
    CampusAnnouncement,
    CampusClass,
    CampusKnowledge,
    CampusRagChunk,
    CampusRagTask,
    CampusCourse,
    CampusCourseOffering,
    CampusCourseSchedule,
    CampusDepartment,
    CampusLeave,
    CampusMessage,
    CampusScore,
    CampusScoreAudit,
    CampusStudent,
    CampusTeacher,
    CampusTerm,
)
from .renderers import _extract_message

User = get_user_model()
from .serializers import (
    AnnouncementSerializer,
    ClassSerializer,
    KnowledgeSerializer,
    CourseSerializer,
    DepartmentSerializer,
    LeaveSerializer,
    OfferingSerializer,
    ScheduleSerializer,
    ScoreAuditSerializer,
    ScoreSerializer,
    StudentSerializer,
    TeacherSerializer,
    TermSerializer,
)


# ===== 权限（6.4/P1-10：仅 admin 角色）=====

class IsAdminRole:
    """DRF 权限类：仅 role_code=admin 可访问 /admin/api/**。"""

    def has_permission(self, request, view) -> bool:
        """视图级权限：仅登录且 role_code=admin 可访问（6.4/P1-10）。"""
        user = request.user
        return bool(user and user.is_authenticated and user.role_code == "admin")

    def has_object_permission(self, request, view, obj) -> bool:
        """对象级权限：queryset 已按逻辑删除过滤，admin 角色即可访问。"""
        # 对象级权限：queryset 已按逻辑删除过滤，admin 角色即可访问
        return self.has_permission(request, view)


# ===== 统一异常处理（6.4：返回格式与应用端一致）=====

class ScheduleConflictError(Exception):
    """排课冲突（4.1 P0-06）：班级/教师时间重叠，转 4091。"""

    def __init__(self, message: str):
        super().__init__(message)


def _validation_codes(detail) -> set:
    """收集 ValidationError detail 中的所有错误码（DRF ErrorDetail.code）。"""
    codes: set = set()

    def _walk(node):
        """递归遍历 DRF detail 树，收集 ErrorDetail.code。"""
        if isinstance(node, dict):
            for v in node.values():
                _walk(v)
        elif isinstance(node, (list, tuple)):
            for v in node:
                if hasattr(v, "code"):
                    codes.add(v.code)
                elif isinstance(v, (dict, list, tuple)):
                    _walk(v)
        elif isinstance(node, ErrorDetail):
            codes.add(node.code)

    _walk(detail)
    return codes


def api_exception_handler(exc, context):
    """DRF 统一异常处理：业务异常返回 HTTP 200 + { code, message, data }（与应用端一致）。"""
    if isinstance(exc, ScheduleConflictError):
        return Response({"code": 4091, "message": str(exc), "data": None}, status=200)
    if isinstance(exc, IntegrityError):
        return Response(
            {"code": 4091, "message": "数据已存在：唯一编码或组合重复，请检查", "data": None},
            status=200,
        )
    if isinstance(exc, ValidationError):
        # 唯一约束（UniqueValidator/UniqueTogetherValidator code='unique'）→ 4091 冲突
        if "unique" in _validation_codes(exc.detail):
            return Response(
                {"code": 4091, "message": "数据已存在：唯一编码或组合重复，请检查", "data": None},
                status=200,
            )
        return Response(
            {"code": 4001, "message": _extract_message(exc.detail), "data": None}, status=200
        )
    if isinstance(exc, NotAuthenticated):
        return Response({"code": 4011, "message": "未登录或登录已过期", "data": None}, status=200)
    if isinstance(exc, PermissionDenied):
        return Response({"code": 4031, "message": "无操作权限", "data": None}, status=200)
    if isinstance(exc, NotFound):
        return Response({"code": 4001, "message": "记录不存在", "data": None}, status=200)
    return drf_exception_handler(exc, context)


# ===== 排课冲突校验与 FOR UPDATE 锁定（4.1 / B-04 / P1-13）=====

def _time_overlap_q(ps: int, pe: int) -> Q:
    """时间段重叠判定：A 与 B 重叠 ⟺ A.start ≤ B.end AND B.start ≤ A.end。"""
    return Q(period_start__lte=pe) & Q(period_end__gte=ps)


def _check_schedule_conflict(term_id, class_id, teacher_id, day_of_week,
                             period_start, period_end, exclude_id=None) -> None:
    """班级冲突 + 教师冲突检查（v1 仅校验这两类，4.1 P0-06）。冲突抛 ScheduleConflictError。"""
    qs = CampusCourseSchedule.objects.filter(
        offering__term_id=term_id,
        day_of_week=day_of_week,
        del_flag="0",
    ).exclude(pk=exclude_id)
    if not qs.exists():
        return

    class_conflict = qs.filter(offering__class_field_id=class_id).filter(
        _time_overlap_q(period_start, period_end)
    ).select_related("offering", "offering__class_field", "offering__course").first()
    if class_conflict is not None:
        off = class_conflict.offering
        raise ScheduleConflictError(
            f"班级冲突：{off.class_field.class_name} 在星期{day_of_week} 第"
            f"{class_conflict.period_start}~{class_conflict.period_end}节已有课程"
            f"《{off.course.course_name}》（第{class_conflict.week_start}~{class_conflict.week_end}周），"
            f"与本次节次 {period_start}~{period_end} 重叠"
        )

    teacher_conflict = qs.filter(offering__teacher_id=teacher_id).filter(
        _time_overlap_q(period_start, period_end)
    ).select_related("offering", "offering__course").first()
    if teacher_conflict is not None:
        off = teacher_conflict.offering
        raise ScheduleConflictError(
            f"教师冲突：该教师星期{day_of_week} 第"
            f"{teacher_conflict.period_start}~{teacher_conflict.period_end}节已有课程"
            f"《{off.course.course_name}》（第{teacher_conflict.week_start}~{teacher_conflict.week_end}周），"
            f"与本次节次 {period_start}~{period_end} 重叠"
        )


def _lock_conflict_offering_rows(term_id, class_id, teacher_id) -> None:
    """FOR UPDATE 锁定涉及的教学班（offering）行（B-04/P1-13，设计原文"锁定涉及的教学班"）。

    按 class_id/teacher_id 数值升序依次加锁：
    1) 班级维度：该学期该班级的全部教学班行（走 uk(term,class,course) 唯一索引）；
    2) 教师维度：该学期该教师的全部教学班行。
    并发排课必须引用已存在的教学班，因此同班/同教师同时段的两个事务
    都会锁到相同的 offering 行 → 互斥串行化；后到事务在锁内重新执行冲突
    检查（当前读）即可发现先到事务已插入的排课（防幻读）。
    所有事务按同一升序加锁，避免加锁顺序死锁。
    """
    for kind, value in sorted((("class", class_id), ("teacher", teacher_id)), key=lambda kv: kv[1]):
        qs = CampusCourseOffering.objects.filter(
            term_id=term_id,
            **({"class_field_id": value} if kind == "class" else {"teacher_id": value}),
        ).order_by("id")
        # 当前读 + 行/间隙锁：必须消费结果集才真正加锁
        list(qs.select_for_update())


# ===== 基类：审计字段 + 逻辑删除 =====

class AdminModelViewSet(viewsets.ModelViewSet):
    """管理端 ViewSet 基类：仅 admin、逻辑删除、自动填充审计字段。"""

    permission_classes = [IsAdminRole]
    filter_backends = [SearchFilter]

    def get_queryset(self):
        """返回未逻辑删除的记录集，并保证稳定排序（分页需要）。"""
        qs = super().get_queryset().filter(del_flag="0")
        if not qs.ordered:  # 分页需稳定排序，避免 UnorderedObjectListWarning
            qs = qs.order_by("id")
        return qs

    def perform_create(self, serializer):
        """创建时自动填充审计字段（create_by/create_time/update_by/update_time/del_flag）。"""
        now = timezone.now()
        uid = self.request.user.id
        serializer.save(create_by=uid, create_time=now,
                        update_by=uid, update_time=now, del_flag="0")

    def perform_update(self, serializer):
        """更新时自动填充 update_by/update_time。"""
        serializer.save(update_by=self.request.user.id, update_time=timezone.now())

    def perform_destroy(self, instance):
        """逻辑删除（5.1：删除走 del_flag，不物理删除被引用记录）。"""
        instance.del_flag = "2"
        instance.update_by = self.request.user.id
        instance.update_time = timezone.now()
        instance.save(update_fields=["del_flag", "update_by", "update_time"])


# ===== T2-1：院系 / 班级 / 课程 / 学期 =====

class DepartmentViewSet(AdminModelViewSet):
    """院系管理（T2-1，/admin/api/departments）。"""

    queryset = CampusDepartment.objects.all()
    serializer_class = DepartmentSerializer
    search_fields = ("dept_name", "dept_code")


class ClassViewSet(AdminModelViewSet):
    """班级管理（T2-1，/admin/api/classes）：指定辅导员/年级/专业/院系。"""

    queryset = CampusClass.objects.select_related("department", "counselor").all()
    serializer_class = ClassSerializer
    search_fields = ("class_name", "class_code")


class CourseViewSet(AdminModelViewSet):
    """课程管理（T2-1，/admin/api/courses）。"""

    queryset = CampusCourse.objects.select_related("department").all()
    serializer_class = CourseSerializer
    search_fields = ("course_name", "course_code")


class TermViewSet(AdminModelViewSet):
    """学期管理（T2-1，/admin/api/terms）：is_current 事务切换（5.1/P1-04）。"""

    queryset = CampusTerm.objects.all()
    serializer_class = TermSerializer
    search_fields = ("term_name",)

    def _set_current(self, obj) -> None:
        """置为当前学期时，先清旧学期（事务内保证任意时刻仅一个）。"""
        if obj.is_current == "1":
            CampusTerm.objects.filter(is_current="1").exclude(pk=obj.pk).update(is_current="0")

    def perform_create(self, serializer):
        """创建学期（事务内保证 is_current 唯一，5.1/P1-04）。"""
        now = timezone.now()
        uid = self.request.user.id
        with transaction.atomic():
            obj = serializer.save(create_by=uid, create_time=now,
                                  update_by=uid, update_time=now, del_flag="0")
            self._set_current(obj)

    def perform_update(self, serializer):
        """更新学期（事务内保证 is_current 唯一，5.1/P1-04）。"""
        with transaction.atomic():
            obj = serializer.save(update_by=self.request.user.id, update_time=timezone.now())
            self._set_current(obj)


# ===== T2-2：教学班 =====

class OfferingViewSet(AdminModelViewSet):
    """教学班管理（T2-2，/admin/api/offerings）：唯一约束 (term, class, course)。

    查询层先做唯一校验给出友好提示；数据库唯一索引兜底（IntegrityError → 4091）。
    """

    queryset = CampusCourseOffering.objects.select_related(
        "course", "term", "class_field", "teacher"
    ).all()
    serializer_class = OfferingSerializer
    search_fields = ("course__course_name", "class_field__class_name")


# ===== T2-3：排课（冲突校验 + FOR UPDATE）=====

class ScheduleViewSet(AdminModelViewSet):
    """排课管理（T2-3，/admin/api/schedules）：班级/教师冲突校验（4.1 P0-06）。"""

    queryset = CampusCourseSchedule.objects.select_related("offering").all()
    serializer_class = ScheduleSerializer
    search_fields = ("offering__course__course_name", "offering__class_field__class_name")

    def _save_with_conflict_check(self, serializer) -> CampusCourseSchedule:
        """冲突校验 + FOR UPDATE 锁定 + 写入同一事务；死锁重试（B-04/P1-13）。"""
        data = serializer.validated_data
        instance = serializer.instance

        offering = data.get("offering") or (instance.offering if instance else None)
        if offering is None:
            raise ValidationError({"offering_id": "教学班不能为空"})

        term_id = offering.term_id
        class_id = offering.class_field_id
        teacher_id = offering.teacher_id
        day = data.get("day_of_week")
        ps, pe = data.get("period_start"), data.get("period_end")
        exclude_id = instance.pk if instance else None

        for attempt in range(3):  # 死锁重试（上限 3 次，B-04）
            try:
                with transaction.atomic():
                    _lock_conflict_offering_rows(term_id, class_id, teacher_id)
                    _check_schedule_conflict(
                        term_id, class_id, teacher_id, day, ps, pe, exclude_id
                    )
                    return self._commit_serializer(serializer)
            except OperationalError as exc:
                # 1213 = deadlock found；重试
                if exc.args and exc.args[0] == 1213 and attempt < 2:
                    continue
                raise

    def _commit_serializer(self, serializer):
        """按创建/更新分别填充审计字段并落库。"""
        now = timezone.now()
        uid = self.request.user.id
        if serializer.instance is None:
            return serializer.save(create_by=uid, create_time=now,
                                   update_by=uid, update_time=now, del_flag="0")
        return serializer.save(update_by=uid, update_time=now)

    def create(self, request, *args, **kwargs):
        """创建排课：冲突校验 + 保存，返回统一响应格式。"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self._save_with_conflict_check(serializer)
        return Response(
            {"code": 0, "message": "排课保存成功", "data": serializer.data}, status=200
        )

    def update(self, request, *args, **kwargs):
        """更新排课：冲突校验（排除自身）+ 保存，返回统一响应格式。"""
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self._save_with_conflict_check(serializer)
        return Response(
            {"code": 0, "message": "排课更新成功", "data": serializer.data}, status=200
        )


# ===== T3-1：公告（状态流转 + RAG 任务联动）=====

class AnnouncementViewSet(AdminModelViewSet):
    """公告管理（T3-1，/admin/api/announcements，4.2/5.3.9）。

    - 状态机：0 草稿 → 1 发布 → 2 下架（4.2，唯一发布方为管理端）；
    - 发布记录 publish_time/publisher_id，发布/下架/删除触发 RAG 任务（8.3）
      与 `ann:version` 缓存失效（P1-11/T3-3）；
    - 院系公告选目标院系（单目标，P1-07；**v2.5/ADR-011：班级公告类型已移除**）。
    """

    queryset = CampusAnnouncement.objects.select_related(
        "target_department", "publisher"
    ).all()
    serializer_class = AnnouncementSerializer
    search_fields = ("title", "content")

    def perform_create(self, serializer):
        """创建公告：自动填充审计字段 + 状态流转联动（发布即写 RAG 任务）。"""
        now = timezone.now()
        uid = self.request.user.id
        with transaction.atomic():
            obj = serializer.save(create_by=uid, create_time=now,
                                  update_by=uid, update_time=now, del_flag="0",
                                  publisher_id=uid)
            on_announcement_saved(obj, old_status="0", operator_id=uid, is_create=True)

    def perform_update(self, serializer):
        """更新公告：状态流转联动（发布/下架/编辑已发布 → RAG 任务 + 缓存失效）。"""
        uid = self.request.user.id
        old_status = serializer.instance.status if serializer.instance else None
        with transaction.atomic():
            obj = serializer.save(update_by=uid, update_time=timezone.now())
            on_announcement_saved(obj, old_status=old_status, operator_id=uid)

    def create(self, request, *args, **kwargs):
        """创建公告：统一响应格式（6.4：管理端返回 { code, message, data }）。"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(
            {"code": 0, "message": "公告创建成功", "data": serializer.data}, status=200
        )

    def destroy(self, request, *args, **kwargs):
        """删除公告：逻辑删除（5.1）后返回统一响应格式（6.4）。"""
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(
            {"code": 0, "message": "公告已删除", "data": {"id": instance.pk}}, status=200
        )

    def perform_destroy(self, instance):
        """逻辑删除公告（5.1）：已发布公告触发 delete RAG 任务 + 缓存失效。"""
        with transaction.atomic():
            self._logical_delete(instance)
            on_announcement_deleted(instance)

    def _logical_delete(self, instance) -> None:
        """逻辑删除落库（del_flag='2'，与 perform_destroy 事务内共用）。"""
        instance.del_flag = "2"
        instance.update_by = self.request.user.id
        instance.update_time = timezone.now()
        instance.save(update_fields=["del_flag", "update_by", "update_time"])

    @staticmethod
    def _do_publish(instance, operator_id: int) -> None:
        """执行发布（草稿/下架 → 发布）：状态 + 发布时间 + RAG 任务。"""
        if instance.status == "1":
            raise ValidationError({"status": "公告已发布，无需重复发布"})
        old_status = instance.status
        instance.status = "1"
        instance.update_by = operator_id
        instance.update_time = timezone.now()
        with transaction.atomic():
            instance.save(update_fields=["status", "update_by", "update_time"])
            on_announcement_saved(instance, old_status=old_status, operator_id=operator_id)

    @staticmethod
    def _do_take_down(instance, operator_id: int) -> None:
        """执行下架（发布 → 下架）：状态 + RAG delete 任务。"""
        if instance.status != "1":
            raise ValidationError({"status": "仅已发布公告可下架"})
        old_status = instance.status
        instance.status = "2"
        instance.update_by = operator_id
        instance.update_time = timezone.now()
        with transaction.atomic():
            instance.save(update_fields=["status", "update_by", "update_time"])
            on_announcement_saved(instance, old_status=old_status, operator_id=operator_id)

    @action(detail=True, methods=["post"], url_path="publish")
    def publish(self, request, pk=None):
        """发布公告（4.2：草稿 → 发布，记录 publish_time）。"""
        instance = self.get_object()
        self._do_publish(instance, request.user.id)
        return Response(
            {"code": 0, "message": "公告已发布", "data": self.get_serializer(instance).data},
            status=200,
        )

    @action(detail=True, methods=["post"], url_path="take-down")
    def take_down(self, request, pk=None):
        """下架公告（4.2：发布 → 下架）。"""
        instance = self.get_object()
        self._do_take_down(instance, request.user.id)
        return Response(
            {"code": 0, "message": "公告已下架", "data": self.get_serializer(instance).data},
            status=200,
        )


# ===== T2-8/T2-9：学生/教师档案管理（方案 B：档案 ↔ sys_user 账号联动）=====

class StudentViewSet(AdminModelViewSet):
    """学生档案管理（T2-8，/admin/api/students，5.3.7）。

    方案 B（用户决策 2026-08-26）：创建档案时自动创建 sys_user
    （username=学号，role_code=student，初始密码默认 123456，姓名写入 nick_name）；
    更新同步学号/姓名；删除（逻辑）同步停用账号（del_flag=2）。
    """

    DEFAULT_PASSWORD = "123456"

    queryset = CampusStudent.objects.select_related("user", "class_field").all()
    serializer_class = StudentSerializer
    search_fields = ("student_no", "user__nick_name", "class_field__class_name")

    def create(self, request, *args, **kwargs):
        """创建档案：统一响应格式（6.4：{ code, message, data }）。"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(
            {"code": 0, "message": "创建成功", "data": serializer.data}, status=200
        )

    def destroy(self, request, *args, **kwargs):
        """删除档案：统一响应格式（6.4），联动停用账号。"""
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(
            {"code": 0, "message": "已删除", "data": {"id": instance.pk}}, status=200
        )

    def _extract_no(self, data) -> str:
        """提取唯一编码（学生：学号 student_no）。"""
        return data["student_no"]

    def perform_create(self, serializer):
        """创建档案：事务内先建 sys_user 账号，再写档案（方案 B）。"""
        uid = self.request.user.id
        now = timezone.now()
        no = self._extract_no(serializer.validated_data)
        with transaction.atomic():
            user = self._create_account(
                no, serializer.validated_data.get("nick_name"),
                serializer.validated_data.get("password"), "student", uid, now,
            )
            serializer.save(user=user, create_by=uid, create_time=now,
                            update_by=uid, update_time=now, del_flag="0")

    def perform_update(self, serializer):
        """更新学生档案：学号/姓名变更同步 sys_user（username/nick_name/student_no）。"""
        uid = self.request.user.id
        instance = serializer.instance
        with transaction.atomic():
            self._sync_account(instance, serializer.validated_data, uid)
            serializer.save(update_by=uid, update_time=timezone.now())

    def perform_destroy(self, instance):
        """逻辑删除档案：同步停用关联账号（del_flag=2 + status=1）。"""
        with transaction.atomic():
            self._logical_delete(instance)
            self._disable_account(instance.user, self.request.user.id)

    def _create_account(self, no, nick_name, password, role, uid, now):
        """创建关联 sys_user（username=学号/工号；冲突 → 4091 明确提示）。"""
        name = nick_name or no
        if User.objects.filter(del_flag="0", username=no).exists():
            raise ValidationError({"student_no": "该学号对应的登录账号已存在"})
        return User.objects.create_user(
            username=no, password=password or self.DEFAULT_PASSWORD, nick_name=name,
            role_code=role, status="0", del_flag="0", student_no=no,
            create_by=uid, create_time=now, update_by=uid, update_time=now,
        )

    def _sync_account(self, instance, data, uid):
        """同步账号：学号变更 → username/student_no；姓名变更 → nick_name。"""
        user = instance.user
        new_no = data.get("student_no")
        if new_no and new_no != user.username:
            if User.objects.filter(del_flag="0", username=new_no).exclude(pk=user.pk).exists():
                raise ValidationError({"student_no": "该学号已被其他账号使用"})
            user.username = new_no
            user.student_no = new_no
        name = data.get("nick_name")
        if name:
            user.nick_name = name
        user.update_by = uid
        user.update_time = timezone.now()
        user.save(update_fields=["username", "nick_name", "student_no", "update_by", "update_time"])

    def _disable_account(self, user, uid):
        """停用关联账号（删除档案联动，5.1 删除走 del_flag）。"""
        if user is None:
            return
        if user.del_flag == "2":
            return
        user.del_flag = "2"
        user.status = "1"
        user.update_by = uid
        user.update_time = timezone.now()
        user.save(update_fields=["del_flag", "status", "update_by", "update_time"])

    def _logical_delete(self, instance) -> None:
        """逻辑删除落库（del_flag='2'）。"""
        instance.del_flag = "2"
        instance.update_by = self.request.user.id
        instance.update_time = timezone.now()
        instance.save(update_fields=["del_flag", "update_by", "update_time"])


class TeacherViewSet(StudentViewSet):
    """教师档案管理（T2-9，/admin/api/teachers，5.3.8，方案 B 联动同学生）。

    角色 teacher、工号唯一；职称/院系写入档案，姓名/密码联动 sys_user。
    **兼任辅导员（ADR-010）**：`is_counselor` + `counselor_class_ids`（1~2 个无辅导员班级）
    同步 `campus_class.counselor_id`；≤2 班由前后端校验，数据库不加约束。
    """

    queryset = CampusTeacher.objects.select_related("user", "department").all()
    serializer_class = TeacherSerializer
    search_fields = ("teacher_no", "user__nick_name", "department__dept_name")

    def _extract_no(self, data) -> str:
        """提取唯一编码（教师：工号 teacher_no）。"""
        return data["teacher_no"]

    def _create_account(self, no, nick_name, password, role, uid, now):
        """教师账号：username=工号（角色 teacher，teacher_no 冗余）。"""
        name = nick_name or no
        if User.objects.filter(del_flag="0", username=no).exists():
            raise ValidationError({"teacher_no": "该工号对应的登录账号已存在"})
        return User.objects.create_user(
            username=no, password=password or self.DEFAULT_PASSWORD, nick_name=name,
            role_code="teacher", status="0", del_flag="0", teacher_no=no,
            create_by=uid, create_time=now, update_by=uid, update_time=now,
        )

    def _sync_account(self, instance, data, uid):
        """教师账号同步：工号变更 → username/teacher_no；姓名变更 → nick_name。"""
        user = instance.user
        new_no = data.get("teacher_no")
        if new_no and new_no != user.username:
            if User.objects.filter(del_flag="0", username=new_no).exclude(pk=user.pk).exists():
                raise ValidationError({"teacher_no": "该工号已被其他账号使用"})
            user.username = new_no
            user.teacher_no = new_no
        name = data.get("nick_name")
        if name:
            user.nick_name = name
        user.update_by = uid
        user.update_time = timezone.now()
        user.save(update_fields=["username", "nick_name", "teacher_no", "update_by", "update_time"])

    # ===== 兼任辅导员（ADR-010）=====

    def perform_create(self, serializer):
        """创建教师：账号联动 + 兼任班级分配（同事务）。"""
        uid = self.request.user.id
        now = timezone.now()
        no = self._extract_no(serializer.validated_data)
        is_counselor = serializer.validated_data.pop("is_counselor", False)
        class_ids = serializer.validated_data.pop("counselor_class_ids", [])
        with transaction.atomic():
            user = self._create_account(
                no, serializer.validated_data.get("nick_name"),
                serializer.validated_data.get("password"), "teacher", uid, now,
            )
            serializer.save(user=user, create_by=uid, create_time=now,
                            update_by=uid, update_time=now, del_flag="0")
            self._apply_counselor_classes(user.id, is_counselor, class_ids, uid)

    def perform_update(self, serializer):
        """更新教师：账号同步 + 兼任班级变更（同事务）。"""
        uid = self.request.user.id
        instance = serializer.instance
        is_counselor = serializer.validated_data.pop("is_counselor", None)
        class_ids = serializer.validated_data.pop("counselor_class_ids", None)
        with transaction.atomic():
            self._sync_account(instance, serializer.validated_data, uid)
            serializer.save(update_by=uid, update_time=timezone.now())
            if is_counselor is not None:
                self._apply_counselor_classes(
                    instance.user_id, is_counselor, class_ids or [], uid
                )

    def perform_destroy(self, instance):
        """删除教师：清空兼任班级指定 + 停用账号（逻辑删除）。"""
        with transaction.atomic():
            now = timezone.now()
            CampusClass.objects.filter(counselor_id=instance.user_id, del_flag="0").update(
                counselor_id=None, update_by=self.request.user.id, update_time=now
            )
            self._logical_delete(instance)
            self._disable_account(instance.user, self.request.user.id)

    def _apply_counselor_classes(self, user_id, is_counselor, class_ids, uid) -> None:
        """应用兼任班级（ADR-010）：清空旧兼任 → 校验（≤2/无占用/有效）→ 写入新指定。

        事务内调用；`counselor_class_ids` 去重后最多 2 个。
        """
        now = timezone.now()
        # 1) 清空该教师当前兼任的班级（编辑/取消兼任场景）
        CampusClass.objects.filter(counselor_id=user_id, del_flag="0").update(
            counselor_id=None, update_by=uid, update_time=now
        )
        if not is_counselor:
            return
        ids = list(dict.fromkeys(class_ids))  # 去重保序
        if not ids:
            raise ValidationError({"counselor_class_ids": "兼任辅导员需选择班级"})
        if len(ids) > 2:
            raise ValidationError({"counselor_class_ids": "每名教师最多兼任 2 个班级"})
        qs = CampusClass.objects.filter(id__in=ids, del_flag="0")
        if qs.count() != len(ids):
            raise ValidationError({"counselor_class_ids": "选择的班级无效或已删除"})
        occupied = qs.exclude(counselor_id__isnull=True).exclude(counselor_id=user_id)
        if occupied.exists():
            raise ValidationError(
                {"counselor_class_ids": f"班级「{occupied.first().class_name}」已有辅导员"}
            )
        # 2) 写入新兼任班级
        qs.update(counselor_id=user_id, update_by=uid, update_time=now)


# ===== M4 成绩管理（T4-1）：查询/Excel 导入/发布/撤销发布/审计 =====

class ScoreViewSet(AdminModelViewSet):
    """成绩管理（T4-1，/admin/api/scores，4.3/5.3.10）。

    - 列表：按教学班/课程/学期查看成绩（只读）；
    - 发布/撤销发布：记录 publish_by/publish_time 并写审计（operation 3/4）；
    - Excel 导入：学号/平时/考试成绩（校验学号归属与分数范围，错误行提示）；
    - 审计查询：/admin/api/score-audits（5.3.11，明细可还原 B-11）。
    """

    queryset = CampusScore.objects.select_related(
        "student__user", "offering__course", "offering__term", "offering__class_field"
    ).all()
    serializer_class = ScoreSerializer

    def get_queryset(self):
        """支持 offering_id / class_id / course_id / term_id 筛选。"""
        qs = super().get_queryset()
        offering_id = self.request.query_params.get("offering_id")
        class_id = self.request.query_params.get("class_id")
        course_id = self.request.query_params.get("course_id")
        term_id = self.request.query_params.get("term_id")
        if offering_id:
            qs = qs.filter(offering_id=offering_id)
        if class_id:
            qs = qs.filter(offering__class_field_id=class_id)
        if course_id:
            qs = qs.filter(offering__course_id=course_id)
        if term_id:
            qs = qs.filter(offering__term_id=term_id)
        return qs

    def _write_audit(self, score, old_detail, new_detail, operation: str) -> None:
        """写成绩审计（5.3.11 / B-11：old/new 明细快照）。"""
        now = timezone.now()
        CampusScoreAudit.objects.create(
            student_id=score.student_id, offering_id=score.offering_id,
            old_score=score.total_score, new_score=score.total_score,
            old_detail=old_detail, new_detail=new_detail,
            operator_id=self.request.user.id, operation=operation,
            operation_time=now, create_by=self.request.user.id,
            create_time=now, update_by=self.request.user.id,
            update_time=now, del_flag="0",
        )

    @action(detail=True, methods=["post"], url_path="publish")
    def publish(self, request, pk=None):
        """发布成绩（4.3：学生端可见；写审计 operation=3）。"""
        obj = self.get_object()
        if obj.is_published == "1":
            raise ValidationError({"is_published": "成绩已发布，无需重复发布"})
        detail = {"usual_score": str(obj.usual_score or ""),
                  "exam_score": str(obj.exam_score or ""),
                  "usual_ratio": obj.usual_ratio, "exam_ratio": obj.exam_ratio}
        obj.is_published = "1"
        obj.publish_by = request.user.id
        obj.publish_time = timezone.now()
        obj.update_by = request.user.id
        obj.update_time = timezone.now()
        with transaction.atomic():
            obj.save(update_fields=["is_published", "publish_by", "publish_time",
                                    "update_by", "update_time"])
            self._write_audit(obj, detail, detail, "3")
        return Response({"code": 0, "message": "成绩已发布",
                         "data": self.get_serializer(obj).data}, status=200)

    @action(detail=True, methods=["post"], url_path="unpublish")
    def unpublish(self, request, pk=None):
        """撤销发布（4.3：学生端不可见；写审计 operation=4）。"""
        obj = self.get_object()
        if obj.is_published == "0":
            raise ValidationError({"is_published": "成绩未发布，无需撤销"})
        detail = {"usual_score": str(obj.usual_score or ""),
                  "exam_score": str(obj.exam_score or ""),
                  "usual_ratio": obj.usual_ratio, "exam_ratio": obj.exam_ratio}
        obj.is_published = "0"
        obj.update_by = request.user.id
        obj.update_time = timezone.now()
        with transaction.atomic():
            obj.save(update_fields=["is_published", "update_by", "update_time"])
            self._write_audit(obj, detail, detail, "4")
        return Response({"code": 0, "message": "成绩已撤销发布",
                         "data": self.get_serializer(obj).data}, status=200)

    @action(detail=False, methods=["post"], url_path="batch-publish")
    def batch_publish(self, request):
        """批量发布成绩（T4-1：ids 列表，事务内逐条发布并写审计 operation=3）。"""
        ids = request.data.get("ids") or []
        if not ids:
            raise ValidationError({"ids": "请选择要发布的成绩"})
        published = skipped = 0
        now = timezone.now()
        uid = request.user.id
        with transaction.atomic():
            for obj in CampusScore.objects.filter(pk__in=ids, del_flag="0"):
                if obj.is_published == "1":
                    skipped += 1
                    continue
                detail = {"usual_score": str(obj.usual_score or ""),
                          "exam_score": str(obj.exam_score or ""),
                          "usual_ratio": obj.usual_ratio, "exam_ratio": obj.exam_ratio}
                obj.is_published = "1"
                obj.publish_by = uid
                obj.publish_time = now
                obj.update_by = uid
                obj.update_time = now
                obj.save(update_fields=["is_published", "publish_by", "publish_time",
                                        "update_by", "update_time"])
                self._write_audit(obj, detail, detail, "3")
                published += 1
        return Response({"code": 0, "message": "批量发布完成",
                         "data": {"published": published, "skipped": skipped}}, status=200)

    @action(detail=False, methods=["post"], url_path="batch-unpublish")
    def batch_unpublish(self, request):
        """批量撤销发布成绩（T4-1：ids 列表，事务内逐条撤销并写审计 operation=4）。"""
        ids = request.data.get("ids") or []
        if not ids:
            raise ValidationError({"ids": "请选择要撤销发布的成绩"})
        unpublished = skipped = 0
        now = timezone.now()
        uid = request.user.id
        with transaction.atomic():
            for obj in CampusScore.objects.filter(pk__in=ids, del_flag="0"):
                if obj.is_published == "0":
                    skipped += 1
                    continue
                detail = {"usual_score": str(obj.usual_score or ""),
                          "exam_score": str(obj.exam_score or ""),
                          "usual_ratio": obj.usual_ratio, "exam_ratio": obj.exam_ratio}
                obj.is_published = "0"
                obj.update_by = uid
                obj.update_time = now
                obj.save(update_fields=["is_published", "update_by", "update_time"])
                self._write_audit(obj, detail, detail, "4")
                unpublished += 1
        return Response({"code": 0, "message": "批量撤销完成",
                         "data": {"unpublished": unpublished, "skipped": skipped}}, status=200)

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """下载成绩导入模板（T4-1：列=班级|学科|学号|平时成绩|考试成绩）。

        前端弹框选定 班级+课程 后调用：定位该教学班（当前学期），自动填入班级全部学生学号，
        表头/边框/列宽与 `docs/成绩导入测试模板.xlsx` 样式一致，平时/考试成绩留空待填。
        """
        from django.http import HttpResponse

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        class_id = request.query_params.get("class_id")
        course_id = request.query_params.get("course_id")
        cls = CampusClass.objects.filter(pk=class_id, del_flag="0").first() if class_id else None
        course = CampusCourse.objects.filter(pk=course_id, del_flag="0").first() if course_id else None
        if cls is None or course is None:
            raise ValidationError({"class_id/course_id": "请选择班级和课程"})

        current_term = CampusTerm.objects.filter(is_current="1", del_flag="0").first()
        offering = None
        if current_term:
            offering = CampusCourseOffering.objects.filter(
                class_field_id=cls.id, course_id=course.id,
                term_id=current_term.id, del_flag="0",
            ).first()
        students = list(CampusStudent.objects.filter(
            class_field_id=cls.id, del_flag="0",
        ).order_by("student_no"))

        wb = Workbook()
        ws = wb.active
        ws.title = "成绩导入"
        headers = ["班级", "学科", "学号", "平时成绩", "考试成绩"]
        ws.append(headers + ["说明（导入时忽略）"])

        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        head_fill = PatternFill("solid", fgColor="D9EAF7")
        for c in ws[1]:
            c.font = Font(bold=True, color="1F4E79")
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.column_dimensions["F"].width = 42
        ws["F1"].font = Font(bold=True, color="1F4E79")

        for i, stu in enumerate(students, start=2):
            ws.append([cls.class_name, course.course_name, stu.student_no, "", ""])
            ws.cell(row=i, column=6, value="").font = Font(color="808080", italic=True, size=9)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
            for c in row:
                c.border = border
                c.alignment = Alignment(horizontal="center")
        for col, w in (("A", 14), ("B", 16), ("C", 14), ("D", 12), ("E", 12)):
            ws.column_dimensions[col].width = w

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="score_import_{cls.class_name}_{course.course_name}.xlsx"'
        )
        wb.save(resp)
        return resp

    @action(detail=False, methods=["post"], url_path="import")
    def import_scores(self, request):
        """Excel 导入成绩（T4-1：模板=班级|学科|学号|平时成绩|考试成绩）。

        - 按（班级, 学科）定位**当前学期**教学班 offering（一份文件可覆盖多班级多学科，
          辅导员兼任多班/每班多学科无需分次上传）；
        - 兼容旧模板：行内班级/学科为空时用前端选择器 class_id/course_id 兜底；
          两者皆缺省时用 offering_id（旧调用方式）；
        - 校验：成绩 0~100、学号属于该班级、（班级,学科）存在当前学期教学班；错误行返回行号+原因。
        """
        file = request.FILES.get("file")
        if not file:
            raise ValidationError({"file": "请上传 Excel 文件"})
        offering_id = request.query_params.get("offering_id") or request.data.get("offering_id")
        class_id = request.query_params.get("class_id") or request.data.get("class_id")
        course_id = request.query_params.get("course_id") or request.data.get("course_id")

        fallback_offering = None
        if offering_id:
            fallback_offering = CampusCourseOffering.objects.filter(
                pk=offering_id, del_flag="0"
            ).select_related("class_field", "course").first()
            if fallback_offering is None:
                raise ValidationError({"offering_id": "教学班不存在"})

        from openpyxl import load_workbook

        try:
            wb = load_workbook(file, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ValidationError({"file": f"Excel 解析失败：{exc}"}) from exc
        ws = wb.active
        ratio = _get_score_ratio_tuple()

        # 按表头自适应列位：新模板「班级|学科|学号|平时成绩|考试成绩」；旧模板「学号|平时成绩|考试成绩」
        header = ["" if v is None else str(v).strip()
                  for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]

        def find_col(*names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        if find_col("班级") is not None:
            i_class, i_course = find_col("班级"), find_col("学科", "课程")
            i_student, i_usual, i_exam = (
                find_col("学号"), find_col("平时"), find_col("考试"),
            )
        else:
            i_class = i_course = None
            i_student = find_col("学号") if find_col("学号") is not None else 0
            i_usual = find_col("平时") if find_col("平时") is not None else 1
            i_exam = find_col("考试") if find_col("考试") is not None else 2

        def cell(row_vals, idx):
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        # 班级/学科 名称↔ID 缓存（支持 名称 或 ID 两种填法）
        class_cache = {c.id: c for c in CampusClass.objects.filter(del_flag="0")}
        class_by_name = {c.class_name: c for c in class_cache.values()}
        course_cache = {c.id: c for c in CampusCourse.objects.filter(del_flag="0")}
        course_by_name = {c.course_name: c for c in course_cache.values()}
        current_term = CampusTerm.objects.filter(is_current="1", del_flag="0").first()
        offering_cache: dict[tuple[int, int], CampusCourseOffering | None] = {}

        def resolve_class(cell):
            if cell is None or str(cell).strip() == "":
                return class_cache.get(int(class_id)) if class_id else None
            v = str(cell).strip()
            if v.isdigit():
                return class_cache.get(int(v))
            return class_by_name.get(v)

        def resolve_course(cell):
            if cell is None or str(cell).strip() == "":
                return course_cache.get(int(course_id)) if course_id else None
            v = str(cell).strip()
            if v.isdigit():
                return course_cache.get(int(v))
            return course_by_name.get(v)

        def resolve_offering(cls, course):
            """(班级, 学科) → 当前学期教学班，结果缓存。"""
            if cls is None or course is None:
                return None
            key = (cls.id, course.id)
            if key in offering_cache:
                return offering_cache[key]
            qs = CampusCourseOffering.objects.filter(
                class_field_id=cls.id, course_id=course.id, del_flag="0"
            )
            if current_term:
                qs = qs.filter(term_id=current_term.id)
            off = qs.select_related("class_field", "course").first()
            offering_cache[key] = off
            return off

        errors: list[dict] = []
        imported = 0
        now = timezone.now()
        uid = request.user.id
        with transaction.atomic():
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                values = list(row)
                class_cell = cell(values, i_class)
                course_cell = cell(values, i_course)
                student_no = cell(values, i_student)
                usual_raw = cell(values, i_usual)
                exam_raw = cell(values, i_exam)
                if all(v in (None, "") for v in (class_cell, course_cell, student_no)):
                    continue  # 空行跳过
                student_no = "" if student_no is None else str(student_no).strip()
                if student_no == "":
                    errors.append({"row": idx, "message": "学号为空"})
                    continue
                try:
                    usual = float(usual_raw)
                    exam = float(exam_raw)
                except (TypeError, ValueError):
                    errors.append({"row": idx, "message": f"学号 {student_no} 成绩格式错误"})
                    continue
                if not (0 <= usual <= 100 and 0 <= exam <= 100):
                    errors.append({"row": idx, "message": f"学号 {student_no} 成绩超出 0~100"})
                    continue
                cls = resolve_class(class_cell)
                course = resolve_course(course_cell)
                # 旧模板兼容：行内班级/学科均未填时，回退到 offering_id 指定的教学班
                if fallback_offering is not None and cls is None and course is None:
                    cls = fallback_offering.class_field
                    course = fallback_offering.course
                if cls is None:
                    errors.append({"row": idx, "message": f"学号 {student_no} 班级不存在或未选择"})
                    continue
                if course is None:
                    errors.append({"row": idx, "message": f"学号 {student_no} 学科不存在或未选择"})
                    continue
                offering = resolve_offering(cls, course)
                if offering is None and fallback_offering is not None:
                    offering = fallback_offering
                if offering is None:
                    errors.append({"row": idx, "message": f"学号 {student_no} 班级「{cls.class_name}」学科「{course.course_name}」无当前学期教学班"})
                    continue
                student = CampusStudent.objects.filter(
                    student_no=student_no, class_field_id=offering.class_field_id, del_flag="0"
                ).first()
                if student is None:
                    errors.append({"row": idx, "message": f"学号 {student_no} 不属于班级「{offering.class_field.class_name}」"})
                    continue
                total = round(usual * ratio[0] / 100 + exam * ratio[1] / 100, 2)
                score, _created = CampusScore.objects.update_or_create(
                    student=student, offering=offering,
                    defaults={
                        "usual_score": usual, "exam_score": exam, "total_score": total,
                        "usual_ratio": ratio[0], "exam_ratio": ratio[1],
                        "is_published": "0", "version": 0,
                        "update_by": uid, "update_time": now, "del_flag": "0",
                    },
                )
                if _created:
                    score.create_by = uid
                    score.create_time = now
                    score.save(update_fields=["create_by", "create_time"])
                imported += 1
        return Response({"code": 0, "message": "导入完成",
                         "data": {"imported": imported, "errors": errors}}, status=200)


def _get_score_ratio_tuple() -> tuple[int, int]:
    """成绩占比字典（campus_score_ratio 首条，格式 '40:60'），缺省 40:60。"""
    from .models import SysDictData

    row = SysDictData.objects.filter(
        dict_type="campus_score_ratio", del_flag="0"
    ).order_by("id").first()
    if row and ":" in (row.dict_value or ""):
        u, e = row.dict_value.split(":")
        return int(u), int(e)
    return 40, 60


class ScoreAuditViewSet(AdminModelViewSet):
    """成绩审计查询（T4-1，/admin/api/score-audits，5.3.11 只读）。"""

    queryset = CampusScoreAudit.objects.select_related(
        "student__user", "offering__course"
    ).all()
    serializer_class = ScoreAuditSerializer
    http_method_names = ["get", "head", "options"]

    def get_queryset(self):
        qs = super().get_queryset()
        student_id = self.request.query_params.get("student_id")
        offering_id = self.request.query_params.get("offering_id")
        class_id = self.request.query_params.get("class_id")
        course_id = self.request.query_params.get("course_id")
        if student_id:
            qs = qs.filter(student_id=student_id)
        if offering_id:
            qs = qs.filter(offering_id=offering_id)
        if class_id:
            qs = qs.filter(offering__class_field_id=class_id)
        if course_id:
            qs = qs.filter(offering__course_id=course_id)
        return qs.order_by("-operation_time")


# ===== M5 请假管理（T5-6）：全部记录 + 管理员干预 =====

class LeaveViewSet(AdminModelViewSet):
    """请假管理（T5-6，/admin/api/leaves，4.6/5.3.12）。

    - 列表：全部记录（按状态/学号筛选）；
    - 干预（P1-15）：管理员强制变更状态，必须填 reason；
      更新状态 + 写站内消息通知学生（记录 old→new/reason/operator/time）。
    """

    queryset = CampusLeave.objects.select_related("student__user").all()
    serializer_class = LeaveSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        status = self.request.query_params.get("status")
        student_no = self.request.query_params.get("student_no")
        if status:
            qs = qs.filter(status=status)
        if student_no:
            qs = qs.filter(student__student_no=student_no)
        return qs.order_by("-create_time")

    @action(detail=True, methods=["post"], url_path="intervene")
    def intervene(self, request, pk=None):
        """管理员干预状态（P1-15：非终态变更/终态回退必须填 reason）。

        状态变化即要求 reason；更新状态并写站内消息（含干预说明），
        禁止直接 SQL/后台编辑 status（只能走干预接口）。
        """
        obj = self.get_object()
        new_status = str(request.data.get("status", ""))
        reason = str(request.data.get("reason") or "").strip()
        if new_status not in {"0", "1", "2", "3"}:
            raise ValidationError({"status": "无效状态（0待审批 1通过 2驳回 3撤销）"})
        if new_status == obj.status:
            raise ValidationError({"status": "状态未变化，无需干预"})
        if not reason:
            raise ValidationError({"reason": "干预必须填写原因（P1-15）"})

        old_status = obj.status
        now = timezone.now()
        uid = request.user.id
        with transaction.atomic():
            obj.status = new_status
            obj.update_by = uid
            obj.update_time = now
            if new_status in {"1", "2"}:
                obj.approver_id = uid
                obj.approve_time = now
                obj.approve_comment = f"[管理员干预]{reason}"
            obj.save(update_fields=["status", "update_by", "update_time",
                                    "approver_id", "approve_time", "approve_comment"])
            status_names = {"0": "待审批", "1": "通过", "2": "驳回", "3": "撤销"}
            CampusMessage.objects.create(
                user_id=obj.student.user_id, msg_type="2",
                title="请假状态干预",
                content=f"您的请假由「{status_names.get(old_status)}」调整为「{status_names.get(new_status)}」：{reason}",
                business_type="leave", business_id=obj.id,
                is_read="0", create_time=now, del_flag="0",
            )
        return Response({"code": 0, "message": "状态已干预",
                         "data": self.get_serializer(obj).data}, status=200)


class UserOptionsView(APIView):
    """用户下拉选项（自建管理前端需要：班级选辅导员、教学班选教师）。

    GET /admin/api/users/options?role=counselor|teacher
    仅返回 del_flag=0 用户，admin 角色可见（P1-10）。
    """

    permission_classes = [IsAdminRole]

    def get(self, request):
        """按角色返回用户选项 [{id, name, username}]，role 缺省返回全部。"""
        role = request.query_params.get("role")
        qs = User.objects.filter(del_flag="0")
        if role:
            qs = qs.filter(role_code=role)
        data = [
            {"id": u.id, "name": u.nick_name or u.username, "username": u.username}
            for u in qs.order_by("id")
        ]
        return Response({"code": 0, "message": "ok", "data": data}, status=200)


class ClassOptionsView(APIView):
    """班级下拉选项（ADR-010：教师兼任辅导员选择无辅导员班级）。

    GET /admin/api/classes/options?without_counselor=1
    without_counselor=1 → 仅返回当前无辅导员（counselor_id IS NULL）的班级。
    """

    permission_classes = [IsAdminRole]

    def get(self, request):
        """返回班级选项 [{id, class_name}]，可按"无辅导员"过滤。"""
        qs = CampusClass.objects.filter(del_flag="0")
        if request.query_params.get("without_counselor") == "1":
            qs = qs.filter(counselor_id__isnull=True)
        data = [{"id": c.id, "class_name": c.class_name} for c in qs.order_by("id")]
        return Response({"code": 0, "message": "ok", "data": data}, status=200)


# ===== T7-2：知识库管理（8.3 / 5.3.15 / P0-08 / P0-09）=====

class KnowledgeViewSet(AdminModelViewSet):
    """知识库管理（T7-2，/admin/api/knowledge，8.3/5.3.15）。

    - CRUD + 发布（publish）/下架（take-down）action；
    - content_hash 变更检测（P0-09）：保存时服务端对剥离 HTML 后的正文
      （含标题）取 SHA-256；hash 与 status 均未变 → 不写向量化任务；
    - 同事务任务写入（P0-08）：发布/编辑已发布 → upsert（source_type=2）；
      下架/删除 → delete；任务与业务数据一起提交。

    RAG 数据源（8.3）：数据源一公告（T3-1 已联动）+ 数据源二知识库（本接口）。
    """

    queryset = CampusKnowledge.objects.select_related("publisher").all()
    serializer_class = KnowledgeSerializer
    search_fields = ("title", "tags", "content")

    def get_queryset(self):
        """列表：分类/状态筛选（6.4）+ 逻辑删除过滤 + 稳定排序。"""
        qs = super().get_queryset()
        category = self.request.query_params.get("category")
        status = self.request.query_params.get("status")
        if category:
            qs = qs.filter(category=category)
        if status:
            qs = qs.filter(status=status)
        return qs.order_by("-update_time", "-id")

    def _save_with_flow(self, serializer, is_create: bool):
        """保存 + content_hash 计算 + 同事务写 RAG 任务（P0-08/P0-09）。"""
        uid = self.request.user.id
        now = timezone.now()
        instance = serializer.instance
        old_status = instance.status if instance else None
        old_hash = instance.content_hash if instance else None
        audit = ({"create_by": uid, "create_time": now, "update_by": uid,
                  "update_time": now, "del_flag": "0"} if is_create
                 else {"update_by": uid, "update_time": now})
        with transaction.atomic():
            obj = serializer.save(**audit)
            # P0-09：新 hash 取保存后的最终 title/content（支持 PATCH 局部更新）
            new_hash = compute_content_hash(obj.title, obj.content)
            obj.content_hash = new_hash
            update_fields = ["content_hash"]
            if obj.status == "1" and obj.publisher_id is None:
                obj.publisher_id = uid  # 首次发布记录发布人
                update_fields.append("publisher")
            obj.save(update_fields=update_fields)
            on_knowledge_saved(obj, old_status=old_status, old_hash=old_hash,
                               is_create=is_create)
        return obj

    def create(self, request, *args, **kwargs):
        """新增知识文档（默认草稿；直接传 status=1 即创建并发布）。"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj = self._save_with_flow(serializer, is_create=True)
        return Response({"code": 0, "message": "知识文档创建成功", "data": serializer.data},
                        status=200)

    def update(self, request, *args, **kwargs):
        """编辑知识文档（已发布且内容变化 → 重新向量化，P0-09）。"""
        partial = kwargs.pop("partial", False)
        serializer = self.get_serializer(self.get_object(), data=request.data,
                                         partial=partial)
        serializer.is_valid(raise_exception=True)
        self._save_with_flow(serializer, is_create=False)
        return Response({"code": 0, "message": "知识文档更新成功", "data": serializer.data},
                        status=200)

    def destroy(self, request, *args, **kwargs):
        """逻辑删除（5.1）：已发布文档触发 delete RAG 任务。"""
        instance = self.get_object()
        with transaction.atomic():
            instance.del_flag = "2"
            instance.update_by = request.user.id
            instance.update_time = timezone.now()
            instance.save(update_fields=["del_flag", "update_by", "update_time"])
            on_knowledge_deleted(instance)
        return Response({"code": 0, "message": "知识文档已删除", "data": {"id": instance.pk}},
                        status=200)

    @action(detail=True, methods=["post"], url_path="publish")
    def publish(self, request, pk=None):
        """发布（草稿 → 发布）：发布即触发向量化（5.3.15）。"""
        instance = self.get_object()
        if instance.status == "1":
            raise ValidationError({"status": "知识文档已发布，无需重复发布"})
        with transaction.atomic():
            instance.status = "1"
            instance.publisher_id = instance.publisher_id or request.user.id
            instance.update_by = request.user.id
            instance.update_time = timezone.now()
            instance.save(update_fields=["status", "publisher", "update_by", "update_time"])
            on_knowledge_saved(instance, old_status="0", old_hash=instance.content_hash)
        return Response({"code": 0, "message": "知识文档已发布，将自动向量化",
                         "data": {"id": instance.pk}}, status=200)

    @action(detail=True, methods=["post"], url_path="take-down")
    def take_down(self, request, pk=None):
        """下架（发布 → 草稿）：移除对应向量（8.3）。"""
        instance = self.get_object()
        if instance.status != "1":
            raise ValidationError({"status": "仅已发布知识文档可下架"})
        with transaction.atomic():
            instance.status = "0"
            instance.update_by = request.user.id
            instance.update_time = timezone.now()
            instance.save(update_fields=["status", "update_by", "update_time"])
            on_knowledge_saved(instance, old_status="1", old_hash=instance.content_hash)
        return Response({"code": 0, "message": "知识文档已下架，向量将移除",
                         "data": {"id": instance.pk}}, status=200)


# ===== T7-3：RAG 索引管理（8.3 全量重建兜底 / 9.6 C-10 状态）=====

REBUILD_REQUEST_KEY = "rag:rebuild_requested"
REBUILDING_KEY = "rag:rebuilding"


class RagIndexView(APIView):
    """RAG 索引状态（T7-3，GET /admin/api/rag/index）。

    返回 {num_docs(Redis FT.INFO)、chunk_total(MySQL 已向量化分片)、
    rebuilding(重建中标记)、latest_tasks(最近任务状态)}（6.4/8.3）。
    """

    permission_classes = [IsAdminRole]

    def get(self, request):
        try:
            client = get_redis_client()
            info = client.execute_command("FT.INFO", "rag_idx")
            pairs = info if isinstance(info, dict) else dict(zip(info[::2], info[1::2]))
            num_docs = int(pairs.get("num_docs") or pairs.get(b"num_docs") or 0)
            rebuilding = bool(client.exists(REBUILDING_KEY))
        except Exception:  # noqa: BLE001 —— Redis 故障时状态接口不报错，字段置空
            num_docs, rebuilding = None, None

        chunk_total = CampusRagChunk.objects.filter(status="1", del_flag="0").count()
        latest_tasks = list(
            CampusRagTask.objects.filter(del_flag="0")
            .order_by("-id").values("id", "operation", "source_type", "source_id",
                                    "status", "retry_count", "last_error",
                                    "create_time", "update_time")[:5]
        )
        data = {
            "num_docs": num_docs,
            "chunk_total": chunk_total,
            "rebuilding": rebuilding,
            "latest_tasks": latest_tasks,
        }
        return Response({"code": 0, "message": "ok", "data": data}, status=200)


class RagIndexRebuildView(APIView):
    """RAG 索引全量重建（T7-3，POST /admin/api/rag/index/rebuild）。

    实现约定（跨进程协作）：索引 DDL/向量写入均在 FastAPI Worker 进程内
    （`app/rag/worker.py`，RediSearch 客户端与 Embedding 链路唯一实现），
    本接口只置 `rag:rebuild_requested` 标记（TTL 防残留）；Worker 下一轮
    检测到请求后执行：置 `rag:rebuilding` → 清空索引 → 重建 → 对全部已发布
    公告+知识库逐源写 upsert 任务 → 完成后删标记（8.3 兜底流程）。
    """

    permission_classes = [IsAdminRole]

    def post(self, request):
        try:
            client = get_redis_client()
            client.set(REBUILD_REQUEST_KEY, "1", ex=86400)  # TTL 防请求残留
        except Exception as exc:  # noqa: BLE001
            return Response({"code": 5000, "message": f"Redis 不可用，无法发起重建: {exc}",
                             "data": None}, status=200)
        return Response({"code": 0,
                         "message": "重建请求已提交，Worker 将在 30s 内开始执行（8.3 全量重建）",
                         "data": None}, status=200)
